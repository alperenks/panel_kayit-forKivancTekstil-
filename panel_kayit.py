

import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import openpyxl
import os
import threading
import subprocess

last_input_barkod = ""
timeout_seconds = 12
timeout_timer = None

EXCEL_DIR = r"D:\BarkodKayit"
os.makedirs(EXCEL_DIR, exist_ok=True)

def get_excel_path():
    today = datetime.now().strftime("%d_%m_%Y")
    return os.path.join(EXCEL_DIR, f"{today}.xlsx")

def create_excel_if_not_exists(path):
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kayıtlar"
        ws.append(["Tarih-Saat", "Bara Barkodu", "Panel Barkodu"])
        wb.save(path)

def zaman_asimi():
    entry_panel.delete(0, tk.END)
    entry_modul.delete(0, tk.END)
    entry_panel.focus()
    status_label.config(text="Kayıt başarıyla eklendi", fg="green")

def zamanlayici_yeniden_baslat():
    global timeout_timer
    if timeout_timer:
        timeout_timer.cancel()
    timeout_timer = threading.Timer(timeout_seconds, zaman_asimi)
    timeout_timer.start()

def kaydet_otomatik():
    global timeout_timer
    panel = entry_panel.get().strip()
    modul = entry_modul.get().strip()

    if not (panel and modul):
        return

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    path = get_excel_path()
    create_excel_if_not_exists(path)

    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        ws.append([now, panel, modul])
        wb.save(path)

        tree.insert("", "end", values=(now, panel, modul))
        entry_panel.delete(0, tk.END)
        entry_modul.delete(0, tk.END)
        entry_panel.focus()
        status_label.config(text="Kayıt başarıyla eklendi.", fg="green")
        if timeout_timer:
            timeout_timer.cancel()
    except PermissionError:
        status_label.config(text="Excel dosyası açık olabilir!", fg="red")

def entry_panel_enter(event):
    global last_input_barkod
    current_input = entry_panel.get().strip()
    if current_input == last_input_barkod:
        status_label.config(text="Kayıt başarıyla eklendi", fg="green")
        entry_panel.delete(0, tk.END)
        return
    last_input_barkod = current_input
    if current_input:
        entry_modul.focus()
        zamanlayici_yeniden_baslat()

def entry_modul_enter(event):
    global last_input_barkod
    current_input = entry_modul.get().strip()
    if current_input == last_input_barkod:
        status_label.config(text="Kayıt başarıyla eklendi", fg="green")
        entry_modul.delete(0, tk.END)
        return
    last_input_barkod = current_input
    if current_input:
        kaydet_otomatik()

def excel_ac():
    path = get_excel_path()
    subprocess.Popen(['start', '', path], shell=True)

def seciliyi_sil():
    selected = tree.selection()
    if selected:
        for item in selected:
            values = tree.item(item, 'values')
            delete_from_excel(values)
            tree.delete(item)

def secilenleri_sil():
    for item in tree.selection():
        values = tree.item(item, 'values')
        delete_from_excel(values)
        tree.delete(item)

def delete_from_excel(values):
    path = get_excel_path()
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if [cell.value for cell in row[:3]] == list(values):
            ws.delete_rows(row[0].row)
            break
    wb.save(path)

# Arayüz
root = tk.Tk()
root.title("KIVANC ENERJI BARA TAKIP SISTEMI")
root.geometry("900x350")
root.resizable(False, False)

# Sol taraf
frame_left = tk.Frame(root, padx=10, pady=10)
frame_left.pack(side=tk.LEFT, fill=tk.Y)

lbl_panel = tk.Label(frame_left, text="Bara Barkodu:")
lbl_panel.pack()
entry_panel = tk.Entry(frame_left, width=30)
entry_panel.pack()
entry_panel.bind("<Return>", entry_panel_enter)

lbl_modul = tk.Label(frame_left, text="Panel Barkodu:")
lbl_modul.pack(pady=(10,0))
entry_modul = tk.Entry(frame_left, width=30)
entry_modul.pack()
entry_modul.bind("<Return>", entry_modul_enter)

status_label = tk.Label(frame_left, text="", font=("Arial", 10))
status_label.pack(pady=10)

excel_button = tk.Button(frame_left, text="Excel Dosyasını Aç", command=excel_ac)
excel_button.pack()

# Sağ taraf
frame_right = tk.Frame(root, padx=10, pady=10)
frame_right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)





tree = ttk.Treeview(frame_right, columns=("Tarih", "Bara", "Panel"), show="headings")
tree.heading("Tarih", text="Tarih-Saat")
tree.heading("Bara", text="Bara Barkodu")
tree.heading("Panel", text="Panel Barkodu")
tree.pack(fill=tk.BOTH, expand=True)

btn_frame = tk.Frame(frame_right)
btn_frame.pack(pady=5)
tk.Button(btn_frame, text="Seçiliyi Sil", bg="red", fg="white", command=seciliyi_sil).pack(side=tk.LEFT, padx=5)
tk.Button(btn_frame, text="Seçilenleri Sil", bg="red", fg="white", command=secilenleri_sil).pack(side=tk.LEFT)

entry_panel.focus()
root.mainloop()
